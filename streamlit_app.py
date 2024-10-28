import os
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime

# Folder Output
output_dir = "output"
os.makedirs(output_dir, exist_ok=True)

# Fungsi Transformasi Data
def transform_data(input_file_path, output_file_path):
    df_input = pd.read_excel(input_file_path, header=5)
    df_input = df_input.dropna(subset=['Order No'])
    wb = Workbook()
    ws = wb.active

    # Styling
    header_fill = PatternFill(start_color="7CE086", end_color="7CE086", fill_type="solid")
    item_fill = PatternFill(start_color="7BA9E1", end_color="7BA9E1", fill_type="solid")
    expense_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Header
    ws.append(["HEADER", "No Form", "Tgl Pesanan", "No Pelanggan", "No PO", "Alamat", "Kena PPN", "Total Termasuk PPN",
               "Diskon Pesanan (%)", "Diskon Pesanan (Rp)", "Keterangan", "Nama Cabang", "Pengiriman", 
               "Tgl Pengiriman", "FOB", "Syarat Pembayaran"])
    for cell in ws[1]:
        cell.fill = header_fill

    # Isi Konten
    unique_orders = df_input['Order No'].unique()
    for order_no in unique_orders:
        order_data = df_input[df_input['Order No'] == order_no]
        order_date = order_data.iloc[0]["Posted Date"]
        if isinstance(order_date, datetime):
            order_date = order_date.strftime("%d/%m/%Y")
        
        ws.append(["HEADER", order_no, order_date, order_data.iloc[0]["Customer Code"], None, 
                   order_data.iloc[0]["Address"], None, None, None, None, None, None, None, None, None, None])
        ws[f"A{ws.max_row}"].fill = header_fill

        for _, row in order_data.iterrows():
            ws.append(["ITEM", row["Item Code"], row["Item Name"], row["Quantity"], row["UOM"], row["Unit Price"], 
                       row["Discount"], None, None, None, None, None, None, None, None, None])
            ws[f"A{ws.max_row}"].fill = item_fill

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file_path)

# Fungsi Membuat Sales Order XML
def create_sales_order_xml(input_file_path, output_file_path):
    df = pd.read_excel(input_file_path, skiprows=5)
    df.columns = [col.strip() for col in df.columns]

    def safe_str(value):
        return '' if pd.isna(value) else str(value)

    df = df.dropna(subset=['Order No'])
    root = ET.Element("NMEXML", EximID="13", BranchCode="2040822216", ACCOUNTANTCOPYID="")
    transactions = ET.SubElement(root, "TRANSACTIONS", OnError="CONTINUE")
    unique_orders = df['Order No'].unique()

    for order_no in unique_orders:
        sales_order = ET.SubElement(transactions, "SALESORDER", operation="Add", REQUESTID="1")
        ET.SubElement(sales_order, "TRANSACTIONID").text = "" 
        order_df = df[df['Order No'] == order_no]
        key_id_counter = 0

        for _, row in order_df.iterrows():
            item_line = ET.SubElement(sales_order, "ITEMLINE", operation="Add")
            ET.SubElement(item_line, "KeyID").text = str(key_id_counter)
            ET.SubElement(item_line, "ITEMNO").text = safe_str(row.get('Item Code', ''))
            ET.SubElement(item_line, "QUANTITY").text = safe_str(row.get('Quantity', ''))
            ET.SubElement(item_line, "ITEMUNIT").text = safe_str(row.get('UOM', ''))
            ET.SubElement(item_line, "ITEMOVDESC").text = safe_str(row.get('Item Name', ''))
            ET.SubElement(item_line, "UNITPRICE").text = safe_str(row.get('Unit Price', ''))
            key_id_counter += 1

        ET.SubElement(sales_order, "SONO").text = safe_str(order_no)
        ET.SubElement(sales_order, "SODATE").text = safe_str(order_df['Posted Date'].iloc[0] if not order_df.empty else "")
        ET.SubElement(sales_order, "CUSTOMERID").text = safe_str(order_df['Customer Code'].iloc[0] if not order_df.empty else "")
        ET.SubElement(sales_order, "CURRENCYNAME").text = "IDR"
        
    xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="    ")
    with open(output_file_path, "w", encoding="utf-8") as f:
        f.write(xml_str)

# Streamlit UI
st.title("Data Processor")
st.sidebar.header("Menu")

# Sidebar Pilihan
options = ["Sales Order", "Purchase Order"]
choice = st.sidebar.radio("Select a feature", options)

# Upload File
uploaded_files = st.file_uploader("Upload Excel file(s)", type=["xlsx"], accept_multiple_files=True)
if uploaded_files:
    for file in uploaded_files:
        file_name = file.name
        file_path = os.path.join(output_dir, file_name)
        with open(file_path, "wb") as f:
            f.write(file.getbuffer())
        
        # Proses Berdasarkan Pilihan
        if choice == "Sales Order":
            st.write("### Processing Sales Order")
            st.write("Convert Excel to desired format:")
            
            if st.button("Convert to Excel (Accurate Online)"):
                output_file_path = os.path.join(output_dir, f"{os.path.splitext(file_name)[0]}_AO.xlsx")
                transform_data(file_path, output_file_path)
                st.success("Data converted to Excel successfully.")
                st.download_button("Download Excel", output_file_path)
                
            if st.button("Convert to XML (Accurate Desktop)"):
                output_file_path = os.path.join(output_dir, f"{os.path.splitext(file_name)[0]}_AD.xml")
                create_sales_order_xml(file_path, output_file_path)
                st.success("Data converted to XML successfully.")
                st.download_button("Download XML", output_file_path)
